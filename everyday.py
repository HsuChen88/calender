from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date
import re

class Event:
    def __init__(self, name):
        self.name = name  # string
        self.time = []  # string
        
    def GetTime(self, count):   # 取得每部影片的"時間長度"
        wb = load_workbook('./excel/course_arr.xlsx')
        ws = wb.active
        if (self.name == "離散"): start_pos = 1
        elif (self.name == "資結"): start_pos = 4
        elif (self.name == "計概"): start_pos = 8
        else: return self.time
        
        # output = []
        for pos in range(3):
            pos = get_column_letter(start_pos + pos)
            getTime = ws[str(pos) + str(count+1)]
            if (getTime.value == None): break
            else: self.time.append(getTime.value)
        
        for i in range(len(self.time)):
            self.time[i] = re.sub("(?:=TIME\()", "", self.time[i])
            self.time[i] = re.sub("(?:\))", "", self.time[i])
            self.time[i] = re.sub(",", ":", self.time[i])
        
        return self.time        

today = date.today()
dateList = re.split("[-]", str(today))      # ['2022', '07', '02']
month = int(dateList[1])
day = int(dateList[2])

wb = load_workbook('./excel/Calender.xlsx')
ws = wb.active
ws = wb[f'{month}月']

# 取得 Calender 上今天的欄位
pattern = f'{month}月{day}日'
for x in range(1, ws.max_column+1):
    char = get_column_letter(x)
    getdate = ws[char + '1'].value
    if getdate == pattern:
        col_num = x
        break
    col_num = None


if col_num == None:
    print("Today is not on the Calender.")
else:
    task_list = []
    char = get_column_letter(col_num)

    i = 2
    task = ws[char + str(i)].value
    while (task != '-'):    # Calender每天所有事情最後會以'-'結尾
        if task == None:
            i += 1
            task = ws[char + str(i)].value
            continue
        else:  
            task_list.append(task)
            i += 1
            task = ws[char + str(i)].value

    # 讀 count.txt 取得 "是否今日第一次打開" 及 "各項課程的記數"
    count_list = []
    with open("./count.txt", mode="r", encoding="utf-8") as file:
        data  = file.read()
    sp_list = re.split("\n", data)
    for item in sp_list:
        tmp_list = re.split("(?:=)", item)   
        count_list.append(tmp_list)
        
    get_today = int(count_list[0][1])
    DS_count = int(count_list[1][1])
    Discrete_count = int(count_list[2][1])
    CD_count = int(count_list[3][1])

    # 若第一次打開 => 記數+1
    for item in task_list:
        if (get_today != day):
            if item == "資結":
                DS_count += 1
            if item == "離散":
                Discrete_count += 1
            if item == "計概":
                CD_count += 1
                
    # 寫入新的記數到 count.txt
    doc = f'Today = {dateList[2]}\nDS_count = {DS_count}\nDiscrete_count = {Discrete_count}\nCD_count = {CD_count}'
    with open("./count.txt", mode="w", encoding="utf-8") as file:
            file.write(doc)

    # 取得各項行程的"時間長度"存放於 time_list
    count = 0
    time_list = []
    for item in task_list:
        obj = Event(item)
        if (item == "離散"): count = Discrete_count
        elif (item == "資結"): count = DS_count
        elif (item == "計概"): count = CD_count
        time_list.append(obj.GetTime(count))
    
    # output
    task_count =  len(task_list)
    print(f'今天有 {task_count} 項任務!\n')
    for i in range(task_count):
        print(str(i+1) + ". " + task_list[i])
        if time_list[i] == []:
            print("")
        else:
            print(time_list[i])
            print("")        
    print("")

input("Press the <Enter> key on the keyboard to exit.")
