# This file help create items in Calender.xlsx

from openpyxl import *
from openpyxl.utils import get_column_letter

# wb = Workbook() # 建立新excel檔
wb = load_workbook('Calender.xlsx') # 載入現有excel檔
# ws.title = "7月" # 更改工作表名稱
ws = wb["7月"] # 操作工作表: 7月
# wb.create_sheet("8月") # 建立工作表

# 產生日期
ls = []
for date in range(1,32):
    ls.append("8月" + str(date) + "日")
ws.append(ls)    

for col in range(1, 32):
    char = get_column_letter(col)
    ws[char + '2'] = f'第一件事'
    ws[char + '3'] = f'第二件事'
    ws[char + '4'] = f'離散'
    ws[char + '5'] = f'資結'
    ws[char + '6'] = f'計概'

li = []
for date in range(1,32):
    li.append("-")  # 每天以 "-" 行程結束
ws.append(li)


wb.save('Calender.xlsx') # 存檔
